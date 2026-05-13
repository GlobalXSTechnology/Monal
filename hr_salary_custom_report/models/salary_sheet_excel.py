from odoo import models, fields
from datetime import datetime
import io
import base64

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    pass


class SalarySheetReportWizard(models.TransientModel):
    _inherit = 'salary.sheet.report.wizard'

    def action_export_excel(self):
        self.ensure_one()

        report_obj = self.env['report.hr_salary_custom_report.salary_sheet_report_template']
        report_data = report_obj._get_report_data(self)

        output = io.BytesIO()
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Salary Sheet"

        header_font = Font(bold=True, size=9, color="FFFFFF")
        header_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        title_font = Font(bold=True, size=14, color="2c3e50")
        subtitle_font = Font(bold=True, size=12, color="7f8c8d")
        info_font = Font(size=9, color="666666")

        dept_header_font = Font(bold=True, size=10, color="FFFFFF")
        dept_header_fill = PatternFill(start_color="203764", end_color="203764", fill_type="solid")

        stats_font = Font(italic=True, size=9)
        stats_fill = PatternFill(start_color="e9ecef", end_color="e9ecef", fill_type="solid")

        total_font = Font(bold=True, size=9)
        total_fill = PatternFill(start_color="e8f4fd", end_color="e8f4fd", fill_type="solid")

        grand_header_font = Font(bold=True, size=11, color="FFFFFF")
        grand_header_fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
        grand_stats_fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
        grand_total_fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")

        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        thick_border = Border(
            left=Side(style='medium'),
            right=Side(style='medium'),
            top=Side(style='medium'),
            bottom=Side(style='medium')
        )

        center_alignment = Alignment(horizontal="center", vertical="center")
        right_alignment = Alignment(horizontal="right", vertical="center")
        left_alignment = Alignment(horizontal="left", vertical="center")

        current_row = 1

        sheet.merge_cells(f'A{current_row}:AB{current_row}')
        title_cell = sheet[f'A{current_row}']
        title_cell.value = report_data['company_name']
        title_cell.font = title_font
        title_cell.alignment = center_alignment
        current_row += 1

        sheet.merge_cells(f'A{current_row}:AB{current_row}')
        subtitle_cell = sheet[f'A{current_row}']
        subtitle_cell.value = f"Salary Sheet - {report_data['period_name']}"
        subtitle_cell.font = subtitle_font
        subtitle_cell.alignment = center_alignment
        current_row += 1

        sheet.merge_cells(f'A{current_row}:AB{current_row}')
        period_cell = sheet[f'A{current_row}']
        period_cell.value = f"Period: {report_data['from_date']} to {report_data['to_date']} | Printed: {report_data['print_date']}"
        period_cell.font = info_font
        period_cell.alignment = center_alignment
        current_row += 2

        if not report_data.get('has_data', False):
            sheet.merge_cells(f'A{current_row}:AB{current_row}')
            no_data_cell = sheet[f'A{current_row}']
            no_data_cell.value = report_data.get('message', 'No salary data found.')
            no_data_cell.font = Font(color="856404")
            no_data_cell.alignment = center_alignment

            workbook.save(output)
            output.seek(0)

            file_name = f"Salary_Sheet_{report_data['period_name'].replace(' ', '_')}.xlsx"
            attachment = self.env['ir.attachment'].create({
                'name': file_name,
                'type': 'binary',
                'datas': base64.b64encode(output.read()),
                'res_model': self._name,
                'res_id': self.id,
                'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            })

            return {
                'type': 'ir.actions.act_url',
                'url': f'/web/content/{attachment.id}?download=true',
                'target': 'self',
            }

        headers = [
            ('Sr No', 6), ('Emp Code', 10), ('Emp Name', 25), ('Designation', 15),
            ('P', 5), ('A', 5), ('L', 5), ('PL', 5), ('UP', 5),
            ('Basic Salary', 12), ('Earned Salary', 12), ('Out Standing', 12),
            ('Salary Per Day', 11), ('Work Days', 9), ('Encashment', 10),
            ('Salary Days', 10), ('Allow', 10), ('Gross Salary', 12),
            ('Umra Ded', 10), ('Bank AC', 10), ('Food over', 10),
            ('Absenty', 10), ('EOBI', 10), ('Loan Deduct', 10),
            ('Fine/Debt', 10), ('Current Advance', 12), ('Final Bank Salary', 13),
            ('Crock Damage', 11), ('Pre Out Standing', 12), ('Net Salary', 12)
        ]

        for col_idx, (header, width) in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            sheet.column_dimensions[col_letter].width = width

        sr_no = 0
        for dept_data in report_data.get('departments', []):
            sheet.merge_cells(f'A{current_row}:AB{current_row}')
            dept_cell = sheet[f'A{current_row}']
            payslip_emp_count = dept_data['dept_totals'].get('employee_count', 0)
            dept_cell.value = f"Department: {dept_data['department_name']} (Pay Slip Employees: {payslip_emp_count})"
            dept_cell.font = dept_header_font
            dept_cell.fill = dept_header_fill
            dept_cell.alignment = center_alignment
            dept_cell.border = border
            current_row += 1

            stats = dept_data.get('dept_statistics', {})
            sheet.merge_cells(f'A{current_row}:AB{current_row}')
            stats_cell = sheet[f'A{current_row}']
            stats_cell.value = f"Total Employees: {stats.get('total_employees', 0)} | New Joining: {stats.get('new_joinings', 0)} | Resigned: {stats.get('resigned_employees', 0)}"
            stats_cell.font = stats_font
            stats_cell.fill = stats_fill
            stats_cell.alignment = center_alignment
            stats_cell.border = border
            current_row += 1

            for col_idx, (header, _) in enumerate(headers, start=1):
                cell = sheet.cell(row=current_row, column=col_idx)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            current_row += 1

            for emp in dept_data.get('employees', []):
                sr_no += 1

                emp_name = emp.get('emp_name', '')
                if emp.get('emp_status'):
                    emp_name = f"{emp_name} ({emp.get('emp_status')})"

                row_data = [
                    sr_no,
                    emp.get('emp_code', ''),
                    emp_name,
                    emp.get('designation', ''),
                    emp.get('present_days', 0),
                    emp.get('absent_days', 0),
                    emp.get('leave_days', 0),
                    emp.get('pl', 0),
                    emp.get('unpaid_leaves', 0),
                    emp.get('wage', '0'),
                    emp.get('basic_salary', '0'),
                    emp.get('out_standing', '0'),
                    emp.get('salary_day', '0'),
                    emp.get('work_days', 0),
                    emp.get('encashment_days', 0),
                    emp.get('salary_days', 0),
                    emp.get('allowance', '0'),
                    emp.get('gross_salary', '0'),
                    emp.get('umra_dept', '0'),
                    emp.get('bank_ac', '0'),
                    emp.get('food_over', '0'),
                    emp.get('absnty', '0'),
                    emp.get('eobi', '0'),
                    emp.get('loan_deduct', '0'),
                    emp.get('fine_debt', '0'),
                    emp.get('current_accm', '0'),
                    emp.get('final_bank', '0'),
                    emp.get('crockery_deduction', '0'),
                    emp.get('pro_out_standing', '0'),
                    emp.get('net_salary', '0')
                ]

                for col_idx, value in enumerate(row_data, start=1):
                    cell = sheet.cell(row=current_row, column=col_idx)
                    cell.value = value
                    cell.border = border

                    if col_idx == 1:
                        cell.alignment = center_alignment
                    elif col_idx in [2, 3, 4]:
                        cell.alignment = left_alignment
                    elif col_idx in [5, 6, 7, 8, 9, 14, 16]:
                        cell.alignment = center_alignment
                    else:
                        cell.alignment = right_alignment

                current_row += 1

            totals = dept_data.get('dept_totals', {})
            total_row_data = [
                '',
                '',
                f"Department Total - {dept_data['department_name']}",
                '',
                totals.get('present_days', 0),
                totals.get('absent_days', 0),
                totals.get('leave_days', 0),
                totals.get('paid_leaves', 0),
                totals.get('unpaid_leaves', 0),
                totals.get('wage', '0'),
                totals.get('basic_salary', '0'),
                totals.get('out_standing', '0'),
                totals.get('salary_day', '0'),
                totals.get('work_days', 0),
                totals.get('encashment_days', 0),
                totals.get('salary_days', 0),
                totals.get('allowance', '0'),
                totals.get('gross_salary', '0'),
                totals.get('umra_dept', '0'),
                totals.get('bank_ac', '0'),
                totals.get('food_over', '0'),
                totals.get('absnty', '0'),
                totals.get('eobi', '0'),
                totals.get('loan_deduct', '0'),
                totals.get('fine_debt', '0'),
                totals.get('current_accm', '0'),
                totals.get('final_bank', '0'),
                totals.get('crockery_deduction', '0'),
                totals.get('pro_out_standing', '0'),
                totals.get('net_salary', '0')
            ]

            for col_idx, value in enumerate(total_row_data, start=1):
                cell = sheet.cell(row=current_row, column=col_idx)
                cell.value = value
                cell.font = total_font
                cell.fill = total_fill
                cell.border = border

                if col_idx <= 4:
                    cell.alignment = center_alignment
                elif col_idx in [5, 6, 7, 8, 9, 14, 16]:
                    cell.alignment = center_alignment
                else:
                    cell.alignment = right_alignment

            current_row += 2

        grand_totals = report_data.get('grand_totals', {})
        grand_stats = report_data.get('grand_statistics', {})

        sheet.merge_cells(f'A{current_row}:AB{current_row}')
        grand_header = sheet[f'A{current_row}']
        grand_header.value = "GRAND TOTAL STATISTICS"
        grand_header.font = grand_header_font
        grand_header.fill = grand_header_fill
        grand_header.alignment = center_alignment
        grand_header.border = thick_border
        current_row += 1

        sheet.merge_cells(f'A{current_row}:AB{current_row}')
        grand_stats_cell = sheet[f'A{current_row}']
        grand_stats_cell.value = f"Total Employees: {grand_stats.get('total_employees', 0)} | New Joining: {grand_stats.get('new_joinings', 0)} | Resigned: {grand_stats.get('resigned_employees', 0)}"
        grand_stats_cell.font = Font(bold=True, size=9)
        grand_stats_cell.fill = grand_stats_fill
        grand_stats_cell.alignment = center_alignment
        grand_stats_cell.border = border
        current_row += 1

        sheet.merge_cells(f'A{current_row}:AB{current_row}')
        grand_sub_header = sheet[f'A{current_row}']
        grand_sub_header.value = "GRAND TOTAL - ALL DEPARTMENTS"
        grand_sub_header.font = Font(bold=True, size=10)
        grand_sub_header.fill = header_fill
        grand_sub_header.alignment = center_alignment
        grand_sub_header.border = border
        current_row += 1

        grand_headers = [
            'Total Depts', 'Total Emps', 'P', 'A', 'L', 'PL', 'UP',
            'Basic Salary', 'Earned Salary', 'Out Standing', 'Salary Per Day',
            'Work Days', 'Encashment', 'Salary Days', 'Allow', 'Gross Salary',
            'Umra Ded', 'Bank AC', 'Food over', 'Absenty', 'EOBI',
            'Loan Deduct', 'Fine/Debt', 'Current Advance', 'Final Bank Salary',
            'Crock Damage', 'Pre Out Standing', 'Net Salary'
        ]

        for col_idx, header in enumerate(grand_headers, start=1):
            cell = sheet.cell(row=current_row, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        current_row += 1

        grand_total_row = [
            len(report_data.get('departments', [])),
            grand_totals.get('employee_count', 0),
            grand_totals.get('present_days', 0),
            grand_totals.get('absent_days', 0),
            grand_totals.get('leave_days', 0),
            grand_totals.get('paid_leaves', 0),
            grand_totals.get('unpaid_leaves', 0),
            grand_totals.get('wage', '0'),
            grand_totals.get('basic_salary', '0'),
            grand_totals.get('out_standing', '0'),
            grand_totals.get('salary_day', '0'),
            grand_totals.get('work_days', 0),
            grand_totals.get('encashment_days', 0),
            grand_totals.get('salary_days', 0),
            grand_totals.get('allowance', '0'),
            grand_totals.get('gross_salary', '0'),
            grand_totals.get('umra_dept', '0'),
            grand_totals.get('bank_ac', '0'),
            grand_totals.get('food_over', '0'),
            grand_totals.get('absnty', '0'),
            grand_totals.get('eobi', '0'),
            grand_totals.get('loan_deduct', '0'),
            grand_totals.get('fine_debt', '0'),
            grand_totals.get('current_accm', '0'),
            grand_totals.get('final_bank', '0'),
            grand_totals.get('crockery_deduction', '0'),
            grand_totals.get('pro_out_standing', '0'),
            grand_totals.get('net_salary', '0')
        ]

        for col_idx, value in enumerate(grand_total_row, start=1):
            cell = sheet.cell(row=current_row, column=col_idx)
            cell.value = value
            cell.font = Font(bold=True, size=10)
            cell.fill = grand_total_fill
            cell.border = border

            if col_idx in [1, 2, 3, 4, 5, 6, 7, 12, 14]:
                cell.alignment = center_alignment
            else:
                cell.alignment = right_alignment

        workbook.save(output)
        output.seek(0)

        file_name = f"Salary_Sheet_{report_data['period_name'].replace(' ', '_')}.xlsx"
        attachment = self.env['ir.attachment'].create({
            'name': file_name,
            'type': 'binary',
            'datas': base64.b64encode(output.read()),
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })

        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }